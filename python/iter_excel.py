import openpyxl
import os
import sys

def iterate_excel_cells(file_path, sheet_name=None):
    """
    Iterates over each cell in the given Excel sheet and prints its coordinates and value.
    :param file_path: Path to the Excel file (.xlsx)
    :param sheet_name: Optional sheet name. If None, uses the active sheet.
    """
    try:
        # Validate file existence
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        # Validate file extension
        if not file_path.lower().endswith(".xlsx"):
            raise ValueError("Only .xlsx files are supported.")

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Select sheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # Iterate over all rows and columns
        for row in sheet.iter_rows():
            for cell in row:
                print(f"Cell {cell.coordinate}: {cell.value}")

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    # Example usage
    # Replace with your file path and optional sheet name
    excel_file = r"C:\Users\olive\OneDrive\12NEW\Miles_2026.xlsx"
    sheet_to_read = None  # or "Sheet1"

    iterate_excel_cells(excel_file, sheet_to_read)
